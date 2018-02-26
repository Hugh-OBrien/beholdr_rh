# -*- coding: utf-8 -*-
import xml.etree.ElementTree as ET
import requests
import json
from os import path, chdir
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

class review:
    def __init__(self, title="", userName="", date="", text="", rating = 0):
        self.title = title
        self.userName = userName
        self.date = date
        self.text = text
        self.rating = rating

class country:
    #class to hold all the data from one country
    def __init__(self, code, name, reviewList, iTunesID):
        self.code = code
        self.name = name.split('\n')[0]
        self.reviews = reviewList
        self.iTunesID = iTunesID

    def getReviews(self, page = 1, recursion = True):
    #takes a country object and makes a call to itunes to gather all review data from that store.
    #adds review data to country objects


        url = "https://itunes.apple.com/" + \
        self.code + "/rss/customerreviews" + "/page=" + str(page) + \
        "/id=" + str(self.iTunesID) +  "/sortBy=mostRecent/json"

        print url
        resp = requests.get(url, verify = "cacert.pem")
        resp = resp.text
        resp = resp.encode("utf8")

        #tree = ET.fromstring(resp)
        parsed = json.loads(resp)
        parsed = parsed['feed']

        #determine if there are more xml pages to view
        lastPageNum = None
        #loop to get the reviews from the json
        if 'entry' not in parsed.keys(): return

        for e in parsed['entry'][1:]:
            self.reviews.append(review(e['title']['label'],
                                       e['author']['name']['label'],
                                       '',
                                       e['content']['label'],
                                       e['im:rating']['label']))

            # if e.tag == st+"link":
            #     if e.get('rel') == "last" and e.get("href")!="":
            #         lastPage = e.get("href")
            #         temp = lastPage.rsplit('page=')
            #         lastPageNum = int(temp[1][:1])
            #     elif e.get("href")=="":
            #         lastPageNum = page
        for p in parsed['link']:
            p = p['attributes']
            if p['rel'] == 'last':
                split1 = p['href'].split('page=')
                split2 = split1[1].split('/')
                lastPageNum = int(split2[0])

        if lastPageNum > page and recursion == True:
            page += 1
            self.getReviews(page,True)

def report(itunesID, countryFile, book, summarySheet, summaryWorkingRow):
    """makes a list of country objects and generates an excel sheet
    with all the reviews from all itunes stores"""
    failedSearches = []
    countryList = []

    try:
        newPath = path.abspath(path.dirname(__file__))
        chdir(newPath)

    except:
        return "can't open cert file, file missing from app folder"

    for c in countryFile:
        newCountryObject = country(c[0:2],c[3:],[],itunesID)
        countryList.append(newCountryObject)

        try:
            #print  >>sys.stderr, "works"
            newCountryObject.getReviews() #while looping through collect review data
        except Exception, e:
            if str(e) != "no element found: line 1, column 0":
                #surpressing this error to simply skip these countries instead
                return str(e) + " - " + newCountryObject.name
            else:
                failedSearches.append(newCountryObject.name)

    outputPath = newPath + "\itunesReviews.xls"

    noReviews = True
    summaryList = []

    workingRow = 2

    for con in countryList:
        #print con.reviews
        summaryList.append((con.name,len(con.reviews)))

        if(len(con.reviews) > 0):

            if noReviews == True:
                noReviews = False

            workingSheet = book.create_sheet(title = con.name)
            revRow = 2

            workingSheet.cell(row=1,column=1).value = "Title"
            workingSheet.cell(row=1,column=2).value = "Date"
            workingSheet.cell(row=1,column=3).value = "UserName"
            workingSheet.cell(row=1,column=4).value = "Text"
            workingSheet.cell(row=1,column=5).value = "Rating"

            for rev in con.reviews:
                workingSheet.cell(row=revRow,column=1).value = rev.title
                #workingSheet.cell(row=revRow,column=2).value = rev.date
                workingSheet.cell(row=revRow,column=3).value = rev.userName
                workingSheet.cell(row=revRow,column=4).value = rev.text
                workingSheet.cell(row=revRow,column=5).value = int(rev.rating)

                revRow += 1

        workingRow+=1

    workingRow = summaryWorkingRow

    for k in summaryList: 
        summarySheet.cell(row=workingRow,column=1).value = k[0]
        summarySheet.cell(row=workingRow,column=2).value = k[1]
        workingRow += 1

    summaryList.sort(key=lambda tup: tup[1], reverse = True)


    '''
    if noReviews == True:
    
        return "No reviews found. If you believe this is incorrect please check \
        the iTunes id " +itunesID+ "is correct"
    '''    
    book.save(outputPath) 
    return (book)
    '''
    if len(failedSearches) > 0:
        print failedSearches
        return Exception
    '''

def idLookup(name):
    #looks up the name by the itunes search tool and returns the id

    url = "https://itunes.apple.com/search?term=" + name
    resp = requests.get(url)
    resp = json.loads(resp.text)
    ids = []; 
    #ids.append(name)   

    ##later we can add ways to pick between search 
    ##results but as a basic input for now
    try: 
        for r in resp['results']: 
            ids.append([r['trackId'],r['kind'], r['artistName'], r['trackName'],
                        r['artworkUrl60']])
    except:
        ids.append("Not Found")
    
    if len(ids) == 0 and name != None and name != "":
        ids.append("No iTunes results!")
    
    return ids

    
    
