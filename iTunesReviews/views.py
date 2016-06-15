import sys
from .models import ReviewReport
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
from django.template import loader, RequestContext
from django import forms
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from os import path, chdir, remove
import itunesreviewreport
import io, random
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.views.generic import TemplateView

def main(request,ID = ""):

    try:
        ID = request.POST['ID']
    except:
        ID = ""

    countryfile=0;

    #get the country list
    #run through the country list file and generate country objects with the country codes and names
    try:
        newPath = path.abspath(path.dirname(__file__))
        chdir(newPath)
        
    except Exception, e:
        return HttpResponse("INTERNAL ERROR: countryList not opened, file missing from app folder")

    try:
        countryfile = open("countrylist.csv").readlines()
        #parse countryfile into dictionary {Region:[country1,country2...],...}
        countryDict = {}
        for c in countryfile:
            sp = c.split(',')
            try:
                countryDict[sp[2]].append(sp[0]+","+sp[1])
            except:
                countryDict[sp[2]] = [sp[0]+","+sp[1]]

        #return HttpResponse(sorted(countryDict.items()))

        sortedCountryDict = OrderedDict(sorted(countryDict.items()))

    except:
        return HttpResponse("can't open countryList, file missing from app folder")
    
    template = loader.get_template('itunesreviews/main.html')
    context = RequestContext(request, {'country_list' : sortedCountryDict, 'Ident':ID,})
    return HttpResponse(template.render(context))


def progressPage(request):
    try:
        idnumber = request.POST['id']
        countryList = request.POST.getlist('countryList[]')
    except:
        print "Failed itunesReview Search"
        return redirect("../iTunesReviews")
    try:
        seshId = (request.POST['seshId'])
    #Generate session id
    except:
        random.seed()
        seshId = str(random.randrange(10001, 20000, 1))
        #print >>sys.stderr, seshId
    try:
        if(len(countryList) >0):
            if(int(idnumber) < 10000000000 and int(idnumber) >= 100000000):        
                #get the next 4 countries
                workingCountryList=[]
                for x in range(0, 1):
                    try:
                        workingCountryList.append(countryList[0])
                        countryList.pop(0)
                    except:
                        break

                runReport(idnumber, workingCountryList, seshId)
                        
                template = loader.get_template('itunesreviews/progress.html')
                context = RequestContext(request, {'country_list':countryList, 'Ident':idnumber,
                                                           'seshId':seshId,})
                return HttpResponse(template.render(context))
                        
        else:
            #now we're going to return he actual report and delete the internal records
            rep= ReviewReport.objects.get(name=seshId)
            
            newPath = path.abspath(path.dirname(__file__))
            chdir(newPath)
            
            returnBook = load_workbook(rep.book)
            returnBook = save_virtual_workbook(returnBook)
            
            #delete the actual file
            newPath = path.abspath(path.dirname(__file__))
            chdir(newPath)
            remove(rep.book)
            
            #delete database entry
            rep.delete()
            
            #return as attachment
            response = HttpResponse(returnBook ,content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="iTunesReviews.xls"'
            return response
    except:
        return redirect("../iTunesReviews")

def runReport(idnumber, countryList, sessionId):
        
    if(int(idnumber) < 10000000000 and int(idnumber) >= 100000000):
        
        try:
            newPath = path.abspath(path.dirname(__file__))
            chdir(newPath)
            
        except Exception, e:
            return HttpResponse("INTERNAL ERROR: file locations missing")             
            
        try:
            rep = ReviewReport.objects.get(name=sessionId)
            bookAddress = rep.book
            try:
                book = load_workbook(bookAddress)
                summarySheet = book.get_sheet_by_name("Summary")
                #print >>sys.stderr, countryList

                book = itunesreviewreport.report(idnumber,countryList,
                                                 book,summarySheet, rep.counter)
                rep.counter += len(countryList)
                rep.save()
                    
                book.save(rep.book)

                #print >>sys.stderr, rep.counter 

                return(book)
            except Exception, e:
                #print >>sys.stderr, e    
                rep.delete()
                return runReport(idnumber, countryList, sessionId);
        
        
        except ObjectDoesNotExist:
            #print >>sys.stderr, "no exist"

            #create book path database entry. 
            #Counter set to 2 as the first writeable row on the summary page
            rep = ReviewReport(name=sessionId, book = '', counter=2) 
            #create the workbook
            book = Workbook()
            #create the summary page
            summarySheet = book.create_sheet()
            summarySheet.title = "Summary"    
            summarySheet.cell(row=1,column=1).value = "Country"
            summarySheet.cell(row=1,column=2).value = "Reviews"
            #remove the default sheet
            book.remove_sheet(book.get_sheet_by_name("Sheet"))

            #save the book and the db entry
            book.save("tempXls/"+sessionId+".xlsx") 
            rep.book= "tempXls/"+sessionId+".xlsx"
            rep.save()
            return runReport(idnumber, countryList, sessionId)
            
        except MultipleObjectsReturned:
            rep = ReviewReport.objects.filter(name=sessionId)
            rep.delete()
            return runReport(idnumber, countryList, sessionId+'2')
            
    else:
        #pass a list with the next country to do
        return HttpResponse("itunes id error, " +  idnumber +  " not in range")
            
def search(request):
    try:
        name = request.POST['name']
    except:
        name = ""

    template = loader.get_template('itunesreviews/search.html')

    #if there is a search term do the search, otherwise render the empty results page
    if(name != None and name != ""):
        context = RequestContext(request, {'results' : itunesreviewreport.idLookup(name), 'SearchTerm':name,})
        return HttpResponse(template.render(context))
    else:
        context = RequestContext(request, {'SearchTerm':name,})
        return HttpResponse(template.render(context))

#itunesreviewreport.report(903688996)
